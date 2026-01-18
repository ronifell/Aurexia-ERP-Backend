"""
Excel Export routes
Provides endpoints to export various data to Excel format
"""
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database import get_db
from models import (
    User, SalesOrder, SalesOrderItem, Shipment, ShipmentItem,
    ProductionOrder, Customer, PartNumber, QualityInspection
)
from auth import get_current_active_user

router = APIRouter(prefix="/exports", tags=["Exports"])


def style_header(ws, max_col):
    """Apply styling to header row"""
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    header_font = Font(bold=True, size=12, color="000000")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border


def auto_adjust_columns(ws):
    """Auto-adjust column widths based on content"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def create_excel_response(wb: Workbook, filename: str):
    """Create a StreamingResponse with the Excel file"""
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    headers = {
        'Content-Disposition': f'attachment; filename="{filename}"',
        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    }
    
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers=headers
    )


@router.get("/sales-orders")
async def export_sales_orders(
    status: Optional[str] = None,
    customer_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Export sales orders to Excel"""
    query = db.query(SalesOrder)
    
    if customer_id:
        query = query.filter(SalesOrder.customer_id == customer_id)
    if status:
        query = query.filter(SalesOrder.status == status)
    
    orders = query.all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Orders"
    
    # Headers
    headers = [
        "PO Number", "Customer Code", "Customer Name", "Order Date", "Due Date",
        "Status", "Total Items", "Total Quantity", "Total Shipped", "Notes"
    ]
    ws.append(headers)
    style_header(ws, len(headers))
    
    # Data
    for order in orders:
        total_items = len(order.items)
        total_qty = sum(item.quantity for item in order.items)
        total_shipped = sum(item.quantity_shipped or 0 for item in order.items)
        
        ws.append([
            order.po_number,
            order.customer.code if order.customer else "",
            order.customer.name if order.customer else "",
            order.order_date.strftime('%Y-%m-%d'),
            order.due_date.strftime('%Y-%m-%d'),
            order.status,
            total_items,
            total_qty,
            total_shipped,
            order.notes or ""
        ])
    
    # Items sheet
    ws_items = wb.create_sheet("Order Items")
    item_headers = [
        "PO Number", "Customer", "Part Number", "Description",
        "Quantity Ordered", "Quantity Produced", "Quantity Shipped",
        "Unit Price", "Total Price", "Status"
    ]
    ws_items.append(item_headers)
    style_header(ws_items, len(item_headers))
    
    for order in orders:
        for item in order.items:
            ws_items.append([
                order.po_number,
                order.customer.name if order.customer else "",
                item.part_number.part_number if item.part_number else "",
                item.part_number.description if item.part_number else "",
                item.quantity,
                item.quantity_produced or 0,
                item.quantity_shipped or 0,
                float(item.unit_price) if item.unit_price else 0,
                float(item.total_price) if item.total_price else 0,
                item.status
            ])
    
    auto_adjust_columns(ws)
    auto_adjust_columns(ws_items)
    
    filename = f"sales_orders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return create_excel_response(wb, filename)


@router.get("/shipments")
async def export_shipments(
    status: Optional[str] = None,
    customer_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Export shipments to Excel"""
    query = db.query(Shipment)
    
    if customer_id:
        query = query.filter(Shipment.customer_id == customer_id)
    if status:
        query = query.filter(Shipment.status == status)
    
    shipments = query.all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Shipments"
    
    # Headers
    headers = [
        "Shipment Number", "Customer Code", "Customer Name", "Sales Order",
        "Shipment Date", "Status", "Tracking Number", "Total Items", "Notes"
    ]
    ws.append(headers)
    style_header(ws, len(headers))
    
    # Data
    for shipment in shipments:
        ws.append([
            shipment.shipment_number,
            shipment.customer.code if shipment.customer else "",
            shipment.customer.name if shipment.customer else "",
            shipment.sales_order.po_number if shipment.sales_order else "",
            shipment.shipment_date,
            shipment.status,
            shipment.tracking_number or "",
            len(shipment.items),
            shipment.notes or ""
        ])
    
    # Items sheet
    ws_items = wb.create_sheet("Shipment Items")
    item_headers = [
        "Shipment Number", "Customer", "Part Number", "Description",
        "Quantity", "Unit Price", "Production Order"
    ]
    ws_items.append(item_headers)
    style_header(ws_items, len(item_headers))
    
    for shipment in shipments:
        for item in shipment.items:
            ws_items.append([
                shipment.shipment_number,
                shipment.customer.name if shipment.customer else "",
                item.part_number.part_number if item.part_number else "",
                item.part_number.description if item.part_number else "",
                item.quantity,
                float(item.unit_price) if item.unit_price else 0,
                item.production_order.po_number if item.production_order else ""
            ])
    
    auto_adjust_columns(ws)
    auto_adjust_columns(ws_items)
    
    filename = f"shipments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return create_excel_response(wb, filename)


@router.get("/production-orders")
async def export_production_orders(
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Export production orders to Excel"""
    query = db.query(ProductionOrder)
    
    if status:
        query = query.filter(ProductionOrder.status == status)
    
    orders = query.all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Production Orders"
    
    # Headers
    headers = [
        "PO Number", "Part Number", "Description", "Sales Order",
        "Quantity", "Completed", "Scrapped", "Status",
        "Start Date", "Due Date", "Priority"
    ]
    ws.append(headers)
    style_header(ws, len(headers))
    
    # Data
    for order in orders:
        ws.append([
            order.po_number,
            order.part_number.part_number if order.part_number else "",
            order.part_number.description if order.part_number else "",
            order.sales_order.po_number if order.sales_order else "",
            order.quantity,
            order.quantity_completed or 0,
            order.quantity_scrapped or 0,
            order.status,
            order.start_date.strftime('%Y-%m-%d') if order.start_date else "",
            order.due_date.strftime('%Y-%m-%d') if order.due_date else "",
            order.priority
        ])
    
    auto_adjust_columns(ws)
    
    filename = f"production_orders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return create_excel_response(wb, filename)


@router.get("/customers")
async def export_customers(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Export customers to Excel"""
    customers = db.query(Customer).filter(Customer.is_active == True).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"
    
    # Headers
    headers = [
        "Code", "Name", "Address", "Contact Person",
        "Phone", "Email", "Delivery Frequency", "Active"
    ]
    ws.append(headers)
    style_header(ws, len(headers))
    
    # Data
    for customer in customers:
        ws.append([
            customer.code,
            customer.name,
            customer.address or "",
            customer.contact_person or "",
            customer.phone or "",
            customer.email or "",
            customer.delivery_frequency or "",
            "Yes" if customer.is_active else "No"
        ])
    
    auto_adjust_columns(ws)
    
    filename = f"customers_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return create_excel_response(wb, filename)


@router.get("/part-numbers")
async def export_part_numbers(
    customer_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Export part numbers to Excel"""
    query = db.query(PartNumber).filter(PartNumber.is_active == True)
    
    if customer_id:
        query = query.filter(PartNumber.customer_id == customer_id)
    
    parts = query.all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Part Numbers"
    
    # Headers
    headers = [
        "Part Number", "Customer Code", "Customer Name",
        "Description", "Material Type", "Unit Price", "Active"
    ]
    ws.append(headers)
    style_header(ws, len(headers))
    
    # Data
    for part in parts:
        ws.append([
            part.part_number,
            part.customer.code if part.customer else "",
            part.customer.name if part.customer else "",
            part.description or "",
            part.material_type or "",
            float(part.unit_price) if part.unit_price else 0,
            "Yes" if part.is_active else "No"
        ])
    
    auto_adjust_columns(ws)
    
    filename = f"part_numbers_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return create_excel_response(wb, filename)


@router.get("/quality-inspections")
async def export_quality_inspections(
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Export quality inspections to Excel"""
    query = db.query(QualityInspection)
    
    if status:
        query = query.filter(QualityInspection.status == status)
    
    inspections = query.all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Quality Inspections"
    
    # Headers
    headers = [
        "Inspection Number", "Production Order", "Part Number",
        "Inspection Type", "Status", "Result", "Inspector",
        "Inspection Date", "Defects Found", "Notes"
    ]
    ws.append(headers)
    style_header(ws, len(headers))
    
    # Data
    for inspection in inspections:
        ws.append([
            inspection.inspection_number,
            inspection.production_order.po_number if inspection.production_order else "",
            inspection.part_number.part_number if inspection.part_number else "",
            inspection.inspection_type,
            inspection.status,
            inspection.result or "",
            inspection.inspector.full_name if inspection.inspector else "",
            inspection.inspection_date.strftime('%Y-%m-%d') if inspection.inspection_date else "",
            inspection.defects_found or 0,
            inspection.notes or ""
        ])
    
    auto_adjust_columns(ws)
    
    filename = f"quality_inspections_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return create_excel_response(wb, filename)
